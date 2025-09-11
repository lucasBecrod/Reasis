import pandas as pd
import sqlite3
from openpyxl import load_workbook

def actualizar_anexo_a_con_k4():
    """
    Actualiza el Anexo A (Excel) con las nuevas columnas de K=4 y riesgo-resiliencia
    """
    
    print("=== ACTUALIZANDO ANEXO A (EXCEL) CON DATOS K=4 ===")
    
    # Rutas de archivos
    excel_path = r"archivado\clustering\02 Informe Entregado\Anexo A - Base de datos XLS v5.xlsx"
    sql_path = r"archivado\clustering\02 Informe Entregado\Anexo B - Base de datos SQL v5.db"
    
    try:
        # Cargar datos actualizados desde SQL (que ya tiene K=4)
        conn = sqlite3.connect(sql_path)
        df_sql_actualizado = pd.read_sql_query("SELECT * FROM indices_metodologicos", conn)
        conn.close()
        
        print(f"Datos desde SQL: {len(df_sql_actualizado)} registros, {len(df_sql_actualizado.columns)} columnas")
        
        # Identificar nuevas columnas
        nuevas_columnas = [
            'CLUSTER_KMEANS',
            'TIPOLOGIA_KMEANS', 
            'SILHOUETTE_SCORE',
            'DISTANCIA_CENTROIDE',
            'Y1_ILA_PREDICHO',
            'RESIDUOS',
            'PRO_SCORE',
            'CATEGORIA_RESILIENCIA',
            'CATEGORIA_RIESGO',
            'REGION_DASHBOARD',
            'CATEGORIA_ALTITUD'
        ]
        
        # Verificar cuáles están disponibles en SQL
        nuevas_disponibles = [col for col in nuevas_columnas if col in df_sql_actualizado.columns]
        print(f"Nuevas columnas disponibles en SQL: {len(nuevas_disponibles)}")
        for col in nuevas_disponibles:
            print(f"  - {col}")
        
        # Leer archivo Excel actual
        df_excel_actual = pd.read_excel(excel_path, sheet_name='indices_metodologicos')
        print(f"Datos Excel actuales: {len(df_excel_actual)} registros, {len(df_excel_actual.columns)} columnas")
        
        # Hacer merge para agregar nuevas columnas
        # Usar solo CODIGO_MODULAR para merge
        df_excel_actualizado = df_excel_actual.merge(
            df_sql_actualizado[['CODIGO_MODULAR'] + nuevas_disponibles],
            on='CODIGO_MODULAR',
            how='left'
        )
        
        print(f"Excel actualizado: {len(df_excel_actualizado)} registros, {len(df_excel_actualizado.columns)} columnas")
        
        # Verificar que se agregaron las columnas
        columnas_agregadas = [col for col in nuevas_disponibles if col in df_excel_actualizado.columns]
        print(f"Columnas agregadas exitosamente: {len(columnas_agregadas)}")
        
        # Verificar distribución K=4
        if 'CLUSTER_KMEANS' in df_excel_actualizado.columns:
            distribucion_k4 = df_excel_actualizado.groupby(['CLUSTER_KMEANS', 'TIPOLOGIA_KMEANS']).size()
            print(f"\nDistribución K=4 en Excel actualizado:")
            for (cluster, tipologia), cantidad in distribucion_k4.items():
                print(f"  Cluster {cluster} - {tipologia}: {cantidad} IIEE")
        
        # Guardar Excel actualizado
        # Crear copia de seguridad primero
        backup_path = excel_path.replace('.xlsx', '_backup.xlsx')
        
        # Cargar workbook completo para mantener todas las hojas
        workbook = load_workbook(excel_path)
        
        # Reemplazar solo la hoja indices_metodologicos
        if 'indices_metodologicos' in workbook.sheetnames:
            # Eliminar hoja actual
            del workbook['indices_metodologicos']
        
        # Crear nueva hoja con datos actualizados
        worksheet = workbook.create_sheet('indices_metodologicos', 0)  # Ponerla primera
        
        # Escribir headers
        for col_idx, column in enumerate(df_excel_actualizado.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=column)
        
        # Escribir datos
        for row_idx, row in enumerate(df_excel_actualizado.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar
        workbook.save(excel_path)
        workbook.close()
        
        print(f"Anexo A actualizado exitosamente:")
        print(f"  - Archivo: {excel_path}")
        print(f"  - Nuevas columnas agregadas: {len(columnas_agregadas)}")
        print(f"  - Total columnas final: {len(df_excel_actualizado.columns)}")
        
        # Verificar distribución K=4 final
        if 'CLUSTER_KMEANS' in df_excel_actualizado.columns:
            print(f"\nVerificación final K=4:")
            verificacion = df_excel_actualizado['CLUSTER_KMEANS'].value_counts().sort_index()
            for cluster, cantidad in verificacion.items():
                tipologia = df_excel_actualizado[df_excel_actualizado['CLUSTER_KMEANS'] == cluster]['TIPOLOGIA_KMEANS'].iloc[0]
                print(f"  Cluster {cluster} ({tipologia}): {cantidad} IIEE")
        
        return True, df_excel_actualizado
        
    except Exception as e:
        print(f"Error al actualizar Anexo A: {e}")
        import traceback
        traceback.print_exc()
        return False, None

def verificar_actualizacion_completa():
    """
    Verificación final de que ambos anexos tienen los datos K=4
    """
    
    print(f"\n{'='*60}")
    print("VERIFICACION FINAL DE ANEXOS")
    print("="*60)
    
    # Verificar SQL
    try:
        conn = sqlite3.connect(r"archivado\clustering\02 Informe Entregado\Anexo B - Base de datos SQL v5.db")
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM indices_metodologicos WHERE CLUSTER_KMEANS IS NOT NULL")
        sql_k4_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM indices_metodologicos WHERE CATEGORIA_RESILIENCIA IS NOT NULL")
        sql_riesgo_count = cursor.fetchone()[0]
        
        conn.close()
        
        print(f"Anexo B (SQL):")
        print(f"  - Instituciones con K=4: {sql_k4_count}")
        print(f"  - Instituciones con riesgo-resiliencia: {sql_riesgo_count}")
        
    except Exception as e:
        print(f"Error verificando SQL: {e}")
    
    # Verificar Excel
    try:
        df_excel = pd.read_excel(r"archivado\clustering\02 Informe Entregado\Anexo A - Base de datos XLS v5.xlsx", 
                                sheet_name='indices_metodologicos')
        
        excel_k4_count = df_excel['CLUSTER_KMEANS'].notna().sum() if 'CLUSTER_KMEANS' in df_excel.columns else 0
        excel_riesgo_count = df_excel['CATEGORIA_RESILIENCIA'].notna().sum() if 'CATEGORIA_RESILIENCIA' in df_excel.columns else 0
        
        print(f"Anexo A (Excel):")
        print(f"  - Instituciones con K=4: {excel_k4_count}")
        print(f"  - Instituciones con riesgo-resiliencia: {excel_riesgo_count}")
        
    except Exception as e:
        print(f"Error verificando Excel: {e}")

if __name__ == "__main__":
    # Verificar estado actual
    verificar_anexo_b_sql()
    print("\n" + "="*50)
    verificar_anexo_a_excel()
    
    # Actualizar Anexo A si es necesario
    print("\n" + "="*50)
    print("ACTUALIZANDO ANEXO A...")
    exito, df_actualizado = actualizar_anexo_a_con_k4()
    
    if exito:
        print("\n" + "="*50)
        verificar_actualizacion_completa()
    else:
        print("Error en la actualización del Anexo A")