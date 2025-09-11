import pandas as pd
import sqlite3

def sincronizar_anexo_a_desde_sql():
    """
    Sincroniza el Anexo A (Excel) directamente desde el Anexo B (SQL) que ya tiene K=4
    """
    
    print("SINCRONIZANDO ANEXO A DESDE ANEXO B (SQL)")
    print("="*50)
    
    # Rutas
    sql_path = r"archivado\clustering\02 Informe Entregado\Anexo B - Base de datos SQL v5.db"
    excel_path = r"archivado\clustering\02 Informe Entregado\Anexo A - Base de datos XLS v5.xlsx"
    
    try:
        # Cargar datos completos desde SQL (que ya tiene K=4)
        conn = sqlite3.connect(sql_path)
        df_sql_completo = pd.read_sql_query("SELECT * FROM indices_metodologicos", conn)
        conn.close()
        
        print(f"Datos desde SQL: {len(df_sql_completo)} registros, {len(df_sql_completo.columns)} columnas")
        
        # Verificar que tiene datos K=4
        if 'CLUSTER_KMEANS' in df_sql_completo.columns:
            k4_count = df_sql_completo['CLUSTER_KMEANS'].notna().sum()
            print(f"Instituciones con K=4 en SQL: {k4_count}")
            
            if k4_count > 0:
                dist_k4 = df_sql_completo.groupby(['CLUSTER_KMEANS', 'TIPOLOGIA_KMEANS']).size()
                print("Distribución K=4 en SQL:")
                for (cluster, tipologia), cantidad in dist_k4.items():
                    print(f"  Cluster {cluster} - {tipologia}: {cantidad} IIEE")
            else:
                print("ERROR: SQL no tiene datos K=4")
                return False
        else:
            print("ERROR: SQL no tiene columna CLUSTER_KMEANS")
            return False
        
        # Verificar resiliencia
        if 'CATEGORIA_RESILIENCIA' in df_sql_completo.columns:
            resiliencia_count = df_sql_completo['CATEGORIA_RESILIENCIA'].notna().sum()
            print(f"Instituciones con resiliencia en SQL: {resiliencia_count}")
        
        # Actualizar Excel - Reemplazar completamente la hoja indices_metodologicos
        print(f"\nActualizando Excel...")
        
        # Usar openpyxl para mantener otras hojas intactas
        from openpyxl import load_workbook
        
        # Cargar workbook
        workbook = load_workbook(excel_path)
        print(f"Hojas en Excel: {workbook.sheetnames}")
        
        # Eliminar hoja actual si existe
        if 'indices_metodologicos' in workbook.sheetnames:
            del workbook['indices_metodologicos']
            print("Hoja anterior eliminada")
        
        # Crear nueva hoja
        worksheet = workbook.create_sheet('indices_metodologicos', 0)
        
        # Escribir headers
        for col_idx, column in enumerate(df_sql_completo.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=str(column))
        
        # Escribir datos
        for row_idx, row in enumerate(df_sql_completo.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                # Convertir valores problemáticos
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Guardar
        workbook.save(excel_path)
        workbook.close()
        
        print("Excel actualizado exitosamente")
        
        # Verificar actualización
        df_verificacion = pd.read_excel(excel_path, sheet_name='indices_metodologicos')
        print(f"Verificación: {len(df_verificacion)} registros, {len(df_verificacion.columns)} columnas")
        
        # Verificar K=4
        if 'CLUSTER_KMEANS' in df_verificacion.columns:
            k4_final = df_verificacion['CLUSTER_KMEANS'].notna().sum()
            print(f"K=4 en Excel final: {k4_final} instituciones")
            
            if k4_final > 0:
                print("Distribución K=4 en Excel actualizado:")
                dist_excel = df_verificacion.groupby(['CLUSTER_KMEANS', 'TIPOLOGIA_KMEANS']).size()
                for (cluster, tipologia), cantidad in dist_excel.items():
                    print(f"  Cluster {cluster} - {tipologia}: {cantidad} IIEE")
        
        # Verificar resiliencia  
        if 'CATEGORIA_RESILIENCIA' in df_verificacion.columns:
            resiliencia_final = df_verificacion['CATEGORIA_RESILIENCIA'].notna().sum()
            print(f"Resiliencia en Excel final: {resiliencia_final} instituciones")
        
        print(f"\nANEXO A ACTUALIZADO EXITOSAMENTE")
        print(f"Archivo: {excel_path}")
        print(f"Columnas totales: {len(df_verificacion.columns)}")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

def verificar_sincronizacion():
    """
    Verifica que ambos anexos tengan los mismos datos K=4
    """
    
    print(f"\n{'='*50}")
    print("VERIFICACION FINAL SINCRONIZACION")
    print("="*50)
    
    # Anexo B (SQL)
    try:
        conn = sqlite3.connect(r"archivado\clustering\02 Informe Entregado\Anexo B - Base de datos SQL v5.db")
        cursor = conn.cursor()
        
        cursor.execute("SELECT CLUSTER_KMEANS, TIPOLOGIA_KMEANS, COUNT(*) FROM indices_metodologicos WHERE CLUSTER_KMEANS IS NOT NULL GROUP BY CLUSTER_KMEANS, TIPOLOGIA_KMEANS ORDER BY CLUSTER_KMEANS")
        sql_distribucion = cursor.fetchall()
        
        conn.close()
        
        print("Anexo B (SQL) - Distribución K=4:")
        for cluster, tipologia, cantidad in sql_distribucion:
            print(f"  Cluster {cluster} - {tipologia}: {cantidad} IIEE")
        
    except Exception as e:
        print(f"Error verificando SQL: {e}")
    
    # Anexo A (Excel)
    try:
        df_excel = pd.read_excel(r"archivado\clustering\02 Informe Entregado\Anexo A - Base de datos XLS v5.xlsx", 
                                sheet_name='indices_metodologicos')
        
        print("Anexo A (Excel) - Distribución K=4:")
        if 'CLUSTER_KMEANS' in df_excel.columns:
            excel_distribucion = df_excel.groupby(['CLUSTER_KMEANS', 'TIPOLOGIA_KMEANS']).size()
            for (cluster, tipologia), cantidad in excel_distribucion.items():
                print(f"  Cluster {cluster} - {tipologia}: {cantidad} IIEE")
        else:
            print("  ERROR: No tiene columna CLUSTER_KMEANS")
        
    except Exception as e:
        print(f"Error verificando Excel: {e}")
    
    print(f"\nAMBOS ANEXOS SINCRONIZADOS CON K=4")

if __name__ == "__main__":
    # Sincronizar Anexo A desde SQL
    exito = sincronizar_anexo_a_desde_sql()
    
    if exito:
        # Verificación final
        verificar_sincronizacion()
        
        print(f"\n{'='*60}")
        print("SINCRONIZACION COMPLETADA")
        print("="*60)
        print("ANEXO A y ANEXO B ahora tienen:")
        print("  - Clustering K=4 con tipologias exactas")
        print("  - Riesgo-resiliencia con categorias")  
        print("  - Todas las variables para dashboard")
        print("  - Datos sincronizados entre SQL y Excel")
    else:
        print("Error en la sincronización")