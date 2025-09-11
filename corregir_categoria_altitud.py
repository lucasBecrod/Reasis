import pandas as pd
import sqlite3

def corregir_categoria_altitud():
    """
    Corrige la columna CATEGORIA_ALTITUD para usar solo rangos numéricos sin etiquetas culturales
    """
    
    print("=== CORRIGIENDO CATEGORIA_ALTITUD ==")
    
    # Rutas
    sql_path = r"archivado\clustering\02 Informe Entregado\Anexo B - Base de datos SQL v5.db"
    excel_path = r"archivado\clustering\02 Informe Entregado\Anexo A - Base de datos XLS v5.xlsx"
    
    def categorizar_altitud_corregida(alt):
        """Nueva función con solo rangos numéricos"""
        if pd.isna(alt):
            return "No disponible"
        alt = float(alt)
        if alt < 500:
            return "0-500m"
        elif alt < 2000:
            return "500-2000m"
        elif alt < 3500:
            return "2000-3500m"
        else:
            return ">3500m"
    
    try:
        # Actualizar SQL primero
        print("Actualizando Anexo B (SQL)...")
        conn = sqlite3.connect(sql_path)
        
        # Cargar datos
        df_sql = pd.read_sql_query("SELECT * FROM indices_metodologicos", conn)
        print(f"Registros en SQL: {len(df_sql)}")
        
        # Verificar columna actual
        if 'CATEGORIA_ALTITUD' in df_sql.columns:
            print("Categorías actuales:")
            categorias_actuales = df_sql['CATEGORIA_ALTITUD'].value_counts()
            for cat, count in categorias_actuales.items():
                print(f"  {cat}: {count}")
        
        # Aplicar nueva categorización
        df_sql['CATEGORIA_ALTITUD'] = df_sql['ALTITUD_MSNM'].apply(categorizar_altitud_corregida)
        
        # Guardar en SQL
        df_sql.to_sql('indices_metodologicos', conn, if_exists='replace', index=False)
        conn.close()
        
        print("Nuevas categorías en SQL:")
        nuevas_categorias = df_sql['CATEGORIA_ALTITUD'].value_counts()
        for cat, count in nuevas_categorias.items():
            print(f"  {cat}: {count}")
        
        # Actualizar Excel
        print("\nActualizando Anexo A (Excel)...")
        
        from openpyxl import load_workbook
        
        # Cargar workbook
        workbook = load_workbook(excel_path)
        
        # Eliminar hoja actual
        if 'indices_metodologicos' in workbook.sheetnames:
            del workbook['indices_metodologicos']
        
        # Crear nueva hoja con datos corregidos
        worksheet = workbook.create_sheet('indices_metodologicos', 0)
        
        # Escribir headers
        for col_idx, column in enumerate(df_sql.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=str(column))
        
        # Escribir datos
        for row_idx, row in enumerate(df_sql.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Guardar Excel
        workbook.save(excel_path)
        workbook.close()
        
        # Verificar Excel actualizado
        df_excel_verificacion = pd.read_excel(excel_path, sheet_name='indices_metodologicos')
        
        print("Nuevas categorías en Excel:")
        nuevas_excel = df_excel_verificacion['CATEGORIA_ALTITUD'].value_counts()
        for cat, count in nuevas_excel.items():
            print(f"  {cat}: {count}")
        
        print(f"\n✅ CATEGORIA_ALTITUD corregida en ambos anexos")
        print("   - Eliminadas etiquetas culturales")
        print("   - Solo rangos numéricos de altura")
        print("   - Ambos anexos sincronizados")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    corregir_categoria_altitud()